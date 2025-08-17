
# from pyproj import CRS
# print(CRS.from_epsg(3857))

from datetime import date
import json
import requests
import pandas as pd
import numpy as np
from pyproj import Transformer

from SpatialKPIDensity import SpatialKPIDensityPlot      # ← preferred
from RangeDict import LTE_Ranges,RangeDict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import os

# ── ensure each LTE_Ranges[...] is a RangeDict ─────────────────────────
LTE_Ranges = {k: (v if isinstance(v, RangeDict) else RangeDict(v))
              for k, v in LTE_Ranges.items()}
# ────────────────────────────────────────────────────────────────────────


class SSV4G:
    def __init__(
        self,
        siteid: str,
        task_date: date,
        *,
        BASE_URL = "http://127.0.0.1:8000",          # <-- change for Docker / prod
        task_id: int = 0

    ):  
        self.siteid = siteid
        self.task_date = task_date 
        self.BASE_URL = BASE_URL
        self.SSV_URL = f"{self.BASE_URL}/ssv"                     # convenience prefix
        self.task_id = task_id

    def query_api(self, path: str, *, params: dict | None = None,
                  as_df: bool = False, timeout: int = 10):
        url = f"{self.SSV_URL}/{path}"
        r   = requests.get(url, params=params, timeout=timeout)

        if r.status_code != 200:
            raise RuntimeError(f"GET {url} -> {r.status_code}: {r.text}")

        payload = r.json()
        return pd.DataFrame(payload) if as_df else payload

    # ----------------------------------------------------------
    def query_data(self):
        # 1) site info  – path parameter, no query params
        self.overall_data = self.query_api(
            f"get_site_info/{self.siteid}", as_df=True
        )

        # 2) KPI rows   – query params
        self.kpi = self.query_api(
            "get_site_kpi",
            params={"siteid": self.siteid, "date": self.task_date},
            as_df=True,
        )
        
        self.cells: list[str] = self.overall_data["siteid_cellid"].unique().tolist()

        self.all_data = self.query_api(
            "get_all_data_by_list",
            params = {"siteid_cellids": json.dumps(self.cells), "date": "2025-01-01"},
            as_df=True,
        )
        # ───────── DEBUG: palette / value sanity check (remove later) ─────────
        # test_kpi = "rsrp"            # pick any KPI column you care about
        # if test_kpi in self.all_data.columns:
        #     bad_vals = [v for v in self.all_data[test_kpi].dropna()
        #                 if v not in LTE_Ranges[test_kpi]]
        #     if bad_vals:
        #         print(f"[WARN] {len(set(bad_vals))} '{test_kpi}' readings "
        #               f"outside colour map → e.g. {sorted(set(bad_vals))[:6]}")
        # ───────────────────────────────────────────────────────────────────────


    def make_tables(self):
        """
        Generates percentage distribution tables for each KPI per cell.
        Returns: dict[cell][kpi] = 2D list (header + rows)
        """
        import numpy as np

        # Define all bins and labels for KPIs
        kpi_bins_labels = {
            "rsrp": (
                [-np.inf, -141, -120, -110, -100, -95, -90, -80, -70, -44, np.inf],
                [
                    "X < -141", "[-141, -120)", "[-120, -110)", "[-110, -100)",
                    "[-100, -95)", "[-95, -90)", "[-90, -80)", "[-80, -70)",
                    "[-70, -44)", "X ≥ -44"
                ],
            ),
            "rsrq": (
                [-np.inf, -20, -18, -12, -10, -6, -3, np.inf],
                [
                    "X < -20", "[-20, -18)", "[-18, -12)", "[-12, -10)",
                    "[-10, -6)", "[-6, -3)", "X ≥ -3"
                ],
            ),
            "rssinr": (
                [-np.inf, -20, -10, 0, 15, 25, 35, 50, np.inf],
                [
                    "X < -20", "[-20, -10)", "[-10, 0)", "[0, 15)",
                    "[15, 25)", "[25, 35)", "[35, 50)", "X ≥ 50"
                ]
            ),
            "dl_throughput": (
                [-np.inf, 0, 1000, 3000, 5000, 10000, np.inf],
                [
                    "X < 0", "[0, 1000)", "[1000, 3000)", "[3000, 5000)",
                    "[5000, 10000)", "X ≥ 10000"
                ],
            ),
            "ul_throughput_mb": (
                [-np.inf, 0, 1000, 3000, 5000, 10000, np.inf],
                [
                    "X < 0", "[0, 1000)", "[1000, 3000)", "[3000, 5000)",
                    "[5000, 10000)", "X ≥ 10000"
                ],
            ),
        }

        self.tables = {}

        for cell in self.cells:
            cell_data = self.all_data[self.all_data["siteid_cellid"] == cell]
            self.tables[cell] = {}

            for kpi, (bins, labels) in kpi_bins_labels.items():
                # If KPI not in data, skip
                if kpi not in cell_data.columns:
                    continue

                col = cell_data[kpi].dropna()
                total = len(col)
                header = [f"{kpi.upper()} (% Sample)", "Range %"]
                rows = []

                if total == 0:
                    # No data for this KPI, just zero rows
                    for label in labels:
                        rows.append([label, 0])
                else:
                    counts, _ = np.histogram(col, bins=bins)
                    percentages = [round(100 * c / total, 2) for c in counts]
                    for label, percent in zip(labels, percentages):
                        rows.append([label, f"{percent:.2f}", percent])  # 3rd column is raw float


                self.tables[cell][kpi] = [header] + rows

    def make_plots(self, kpi_list=None, *, pad=1.1, min_km=2.0, max_km=8.0):
        """
        Build one SpatialKPIDensity PNG for every <cell, KPI> pair.

        * window radius = farthest GPS distance × pad (10 % default)
        * clamped between min_km and max_km
        * grid_size fixed at 50 m
        """
        import numpy as np
        kpi_list = kpi_list or list(LTE_Ranges.keys())
        self.plots = {}

        for cell in self.cells:
            df = self.all_data[self.all_data["siteid_cellid"] == cell]
            self.plots[cell] = {}
            if df.empty:
                continue

            meta   = self.overall_data[self.overall_data["siteid_cellid"] == cell].iloc[0]
            bs_lat = float(meta["latitude"])
            bs_lon = float(meta["longitude"])           # ← base-station coords are FIXED


            # dy = (df["latitude"]  - bs_lat).abs().max() * 111_000
            # dx = (df["longitude"] - bs_lon).abs().max() *  85_000
            # win_km = np.clip(np.hypot(dx, dy) * pad / 1_000, min_km, max_km)
            
            # one-time initialisation -----------------------------------
            project = Transformer.from_crs("EPSG:4326", "EPSG:3857", always_xy=True).transform

            # ----------------------------------------------------------
            # df already has longitude / latitude columns
            xs, ys = project(df["longitude"].values, df["latitude"].values)

            # base-station projected coords
            bs_x, bs_y = project(bs_lon, bs_lat)

            dx = np.abs(xs - bs_x).max()          # max east-west offset  (m)
            dy = np.abs(ys - bs_y).max()          # max north-south offset (m)

            pad_factor = 1.20                     # 20 % breathing room
            win_km = np.clip(np.hypot(dx, dy) * pad_factor / 1_000, min_km, max_km)

            for kpi in kpi_list:
                if kpi not in df.columns:
                    continue

                plotter = SpatialKPIDensityPlot(
                    bs_lat=bs_lat, bs_lon=bs_lon,
                    azimuth=float(meta.get("azimuth", 0)),
                    beamwidth=float(meta.get("beamwidth", 60)),
                    radius=100,
                    grid_size=50,
                    data_points=df,
                    lon_col="longitude", lat_col="latitude",
                    kpi_col=kpi, kpi_name=kpi.upper(),
                    kpi_range_dict=LTE_Ranges[kpi],
                    extent_km=round(float(win_km), 1),
                )
                self.plots[cell][kpi] = plotter.plot(out=None)

        return self.plots





    def section_header(self, ws, title, row, col=2, span=18):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
        cell = ws.cell(row=row, column=col)
        cell.value = title
        cell.font = Font(bold=True, color="FFFFFF", size=16)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="3E82FC")
        return row + 2

    def _add_image_block(self, ws, img: XLImage, row, col, n_cols, n_rows):
        img.anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(img)


    def _add_table_block(self, ws, table, row, col, fit_columns=False):
        border = Border(left=Side(style='thin', color='BBBBBB'),
                        right=Side(style='thin', color='BBBBBB'),
                        top=Side(style='thin', color='BBBBBB'),
                        bottom=Side(style='thin', color='BBBBBB'))
        header_fill = PatternFill("solid", fgColor="D1E6FA")
        band1 = PatternFill("solid", fgColor="F4F8FB")
        band2 = PatternFill("solid", fgColor="E9F1F7")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for i, row_data in enumerate(table):
            for j, val in enumerate(row_data):
                c = ws.cell(row=row + i, column=col + j, value=val)
                c.border = border
                c.alignment = center
                if i == 0:
                    c.fill = header_fill
                    c.font = Font(bold=True, name="Calibri", size=11)
                elif i % 2 == 0:
                    c.fill = band1
                else:
                    c.fill = band2

        if fit_columns:
            self.fit_table_columns(ws, col, table)
        return row + len(table) - 1

    def fit_table_columns(self, ws, start_col, table):
        max_lens = [max(len(str(row[i])) for row in table) for i in range(len(table[0]))]
        for i, l in enumerate(max_lens):
            col_letter = get_column_letter(start_col + i)
            ws.column_dimensions[col_letter].width = max(10, min(30, l + 4))

    # ===============================
    # 1) write_report_onepage
    # ===============================
    def write_report_onepage(self, out_xlsx_path: str | None = None, site_name: str | None = None):
        """
        Build a single‑sheet Excel “one‑pager”.
        * Overview tables at the top
        * For every **cell / KPI** block:
              map  →  2‑column table  →  single‑series bar chart
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "SSV 4G Report"
        ws.sheet_view.showGridLines = False

        # --- base column / row sizing –––
        for c in range(2, 30):
            ws.column_dimensions[get_column_letter(c)].width = 24
        for r in range(1, 400):
            ws.row_dimensions[r].height = 18

        cur = 2
        cur = self.section_header(ws, f"4G Site Report: {site_name or self.siteid}", cur) + 2

        # ── 1.  OVERVIEW ─────────────────────────────────────────────
        cur = self.section_header(ws, "Overview - Site Info", cur)
        overview_tbl = [list(self.overall_data.columns)] + self.overall_data.values.tolist()
        cur = self._add_table_block(ws, overview_tbl, cur, 2, fit_columns=True) + 3

        cur = self.section_header(ws, "Overview - KPI Summary", cur)
        kpi_tbl = [list(self.kpi.columns)] + self.kpi.values.tolist()
        cur = self._add_table_block(ws, kpi_tbl, cur, 2, fit_columns=True) + 3

        # ── 2.  COVERAGE & PERFORMANCE (loop every cell / KPI) ───────
        cur = self.section_header(ws, "Coverage and Performances", cur) + 1

        for cell, kpi_dict in self.tables.items():
            ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=8)
            ws.cell(row=cur, column=2, value=f"Cell: {cell}").font = Font(bold=True, color="3E82FC")
            cur += 1

            for kpi, source_tbl in kpi_dict.items():
                ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=8)
                ws.cell(row=cur, column=2, value=kpi.upper()).font = Font(bold=True)
                cur += 1

                                # (a)  map — only if make_plots() actually produced one
                if kpi in self.plots.get(cell, {}):
                    self._add_image_block(ws, self.plots[cell][kpi], cur, 2, n_cols=6, n_rows=15)
                else:
                    # no plot → push table/chart leftwards by 6 cols
                    tbl_col_offset = -6   # shrink layout
                    ws.row_dimensions[cur].height = 12
                
                # (b)  build trimmed 2‑col table  [Range | %] build trimmed 2‑col table  [Range | %]
                tbl_col = 2 + 6 + 1
                trimmed = [["Range", "%"]]
                for row in source_tbl[1:]:               # skip header
                    if len(row) < 3:
                        continue
                    try:
                        pct = float(row[2])
                    except (TypeError, ValueError):
                        continue
                    trimmed.append([str(row[0]).strip(), pct])

                tbl_end = self._add_table_block(ws, trimmed, cur, tbl_col, fit_columns=True)

                # (c)  bar chart – only if we have at least 1 data row > 0
                if len(trimmed) > 1 and any(r[1] > 0 for r in trimmed[1:]):
                    chart_col = tbl_col + 3
                    self._add_chart_block(
                        ws, cur, chart_col,
                        cur, tbl_col, len(trimmed),
                        title=f"{kpi.upper()} Distribution"
                    )

                # move cursor below image/table/chart cluster
                block_bottom = max(cur + 15, tbl_end)   # 15 rows for the image area
                cur = block_bottom + 12                  # breathing-room (tweak to taste)
                ws.row_dimensions[cur].height = 12
        if not out_xlsx_path:                               # None, "" or False
            # make sure self.task_date is a date-like object or ISO string
            date_str = (self.task_date.strftime("%Y-%m-%d")      # datetime/date
                        if hasattr(self.task_date, "strftime")
                        else str(self.task_date))                # already a str
            out_xlsx_path = f"{self.siteid}_{date_str}_4G.xlsx"
            if not os.path.exists("outputs\\" + str(self.task_id) ):
                os.makedirs("outputs\\" + str(self.task_id))
            wb.save("outputs\\" + str(self.task_id)+"\\"+out_xlsx_path)
        else:
            wb.save(out_xlsx_path or "SSV_4G_Report.xlsx")

    # ------------------------------------------------------------------
    def _add_chart_block(
        self,
        ws,
        row: int,
        col: int,
        table_row: int,
        table_col: int,
        table_len: int,
        title: str,
    ):
        """Insert a *single‑series* vertical bar chart that uses the 2‑column
        table starting at *(table_row, table_col)* as its data source."""
        first = table_row + 1  # first data row
        last = table_row + table_len - 1
        if last < first:
            return  # nothing to plot

        # Build chart references
        cats = Reference(ws, min_col=table_col, min_row=first, max_row=last)
        vals = Reference(ws, min_col=table_col + 1, min_row=first, max_row=last)

        chart = BarChart()
        chart.type = "col"
        chart.shape = 4
        chart.height = 11
        chart.width = 20
        chart.add_data(vals, titles_from_data=False, from_rows=False)  # single series
        chart.set_categories(cats)

        # ── axes ────────────────────────────────────────────────
        chart.x_axis.title = None
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.textRotation = 45
        chart.x_axis.majorTickMark = "out"
        chart.x_axis.majorGridlines = None
        chart.x_axis.delete = False  # make sure axis is visible

        chart.y_axis.title = None
        chart.y_axis.majorGridlines = None
        chart.y_axis.delete = False

        # Add ≈20 % head‑room so the tallest bar doesn’t hit the title
        val_col = table_col + 1
        max_val = 0
        for r in ws.iter_rows(min_col=val_col, max_col=val_col, min_row=first, max_row=last):
            try:
                v = float(r[0].value)
            except (TypeError, ValueError):
                v = 0
            max_val = max(max_val, v)
        if max_val > 0:
            chart.y_axis.scaling.max = max_val * 1.2

        # ── labels & legend ─────────────────────────────────────
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True      # numeric value on bars
        chart.dataLabels.showSerName = False # hide “Series1” etc.
        chart.dataLabels.showCatName = False # keep cat names only on x‑axis
        chart.legend = None                  # hide redundant legend

        # ── title ───────────────────────────────────────────────
        chart.title = title
        chart.title_overlay = False

        # finally add chart at top‑left of the chart region
        ws.add_chart(chart, f"{get_column_letter(col)}{row}")




    def __str__(self):
        return ", \n".join(f"{k}=\n{v}\n" for k, v in vars(self).items())

if __name__ == "__main__":
    from timeit import default_timer as timer

    start = timer()
    report = SSV4G(siteid="69491",task_date="2025-01-01")
    report.query_data()
    report.make_tables()
    report.make_plots()
    report.write_report_onepage()
    # print(str(report))
    end = timer()
    print(end - start) # Time in seconds, e.g. 5.38091952400282

