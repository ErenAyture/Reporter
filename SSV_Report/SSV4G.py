
# from pyproj import CRS
# print(CRS.from_epsg(3857))

from datetime import date
import json
import requests
import pandas as pd
import numpy as np
from SpatialKPIDensity import SpatialKPIDensityPlot      # ← preferred
from Grid50Plot import Grid50Plot
from RangeDict import LTE_Ranges

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList


class SSV4G:
    def __init__(
        self,
        siteid: str,
        task_date: date,
        *,
        BASE_URL = "http://127.0.0.1:8000"          # <-- change for Docker / prod

    ):  
        self.siteid = siteid
        self.task_date = task_date 
        self.BASE_URL = BASE_URL
        self.SSV_URL = f"{self.BASE_URL}/ssv"                     # convenience prefix

    def query_api(self, path: str, *, params: dict | None = None,
                  as_df: bool = False, timeout: int = 10):
        url = f"{self.SSV_URL}/{path}"
        r   = requests.get(url, params=params, timeout=timeout)

        if r.status_code != 200:
            raise RuntimeError(f"GET {url} -> {r.status_code}: {r.text}")

        payload = r.json()
        return pd.DataFrame(payload) if as_df else payload

    # ----------------------------------------------------------
    def query_data(self):
        # 1) site info  – path parameter, no query params
        self.overall_data = self.query_api(
            f"get_site_info/{self.siteid}", as_df=True
        )

        # 2) KPI rows   – query params
        self.kpi = self.query_api(
            "get_site_kpi",
            params={"siteid": self.siteid, "date": self.task_date},
            as_df=True,
        )
        
        self.cells: list[str] = self.overall_data["siteid_cellid"].unique().tolist()

        self.all_data = self.query_api(
            "get_all_data_by_list",
            params = {"siteid_cellids": json.dumps(self.cells), "date": "2025-01-01"},
            as_df=True,
        )
    def make_tables(self):
        """
        Generates percentage distribution tables for each KPI per cell.
        Returns: dict[cell][kpi] = 2D list (header + rows)
        """
        import numpy as np

        # Define all bins and labels for KPIs
        kpi_bins_labels = {
            "rsrp": (
                [-np.inf, -141, -120, -110, -100, -95, -90, -80, -70, -44, np.inf],
                [
                    "X < -141", "[-141, -120)", "[-120, -110)", "[-110, -100)",
                    "[-100, -95)", "[-95, -90)", "[-90, -80)", "[-80, -70)",
                    "[-70, -44)", "X ≥ -44"
                ],
            ),
            "rsrq": (
                [-np.inf, -20, -18, -12, -10, -6, -3, np.inf],
                [
                    "X < -20", "[-20, -18)", "[-18, -12)", "[-12, -10)",
                    "[-10, -6)", "[-6, -3)", "X ≥ -3"
                ],
            ),
            "rssinr": (
                [-np.inf, -20, -10, 0, 15, 25, 35, 50, np.inf],
                [
                    "X < -20", "[-20, -10)", "[-10, 0)", "[0, 15)",
                    "[15, 25)", "[25, 35)", "[35, 50)", "X ≥ 50"
                ]
            ),
            "dl_throughput": (
                [-np.inf, 0, 1000, 3000, 5000, 10000, np.inf],
                [
                    "X < 0", "[0, 1000)", "[1000, 3000)", "[3000, 5000)",
                    "[5000, 10000)", "X ≥ 10000"
                ],
            ),
            "ul_throughput_mb": (
                [-np.inf, 0, 1000, 3000, 5000, 10000, np.inf],
                [
                    "X < 0", "[0, 1000)", "[1000, 3000)", "[3000, 5000)",
                    "[5000, 10000)", "X ≥ 10000"
                ],
            ),
        }

        self.tables = {}

        for cell in self.cells:
            cell_data = self.all_data[self.all_data["siteid_cellid"] == cell]
            self.tables[cell] = {}

            for kpi, (bins, labels) in kpi_bins_labels.items():
                # If KPI not in data, skip
                if kpi not in cell_data.columns:
                    continue

                col = cell_data[kpi].dropna()
                total = len(col)
                header = [f"{kpi.upper()} (% Sample)", "Range %"]
                rows = []

                if total == 0:
                    # No data for this KPI, just zero rows
                    for label in labels:
                        rows.append([label, 0])
                else:
                    counts, _ = np.histogram(col, bins=bins)
                    percentages = [round(100 * c / total, 2) for c in counts]
                    for label, percent in zip(labels, percentages):
                        rows.append([label, percent])

                self.tables[cell][kpi] = [header] + rows

    # def make_plots(self, kpi_list=None):
    #     """
    #     Generate plots for all cells and all (or given) KPIs.
    #     Returns: dict[cell][kpi] = image (openpyxl Image)
    #     """

    #     self.plots = {}
    #     kpi_list = list(LTE_Ranges.keys())  # All KPIs by default

    #     # Iterate all cells
    #     for cell in self.cells:
    #         # Row for current cell to get azimuth/beamwidth and lat/lon
    #         row = self.overall_data.loc[self.overall_data["siteid_cellid"] == cell].iloc[0]
    #         cell_data = self.all_data[self.all_data["siteid_cellid"] == cell]
    #         self.plots[cell] = {}

    #         for kpi in kpi_list:
    #             # Check if KPI exists in data (optional)
    #             if kpi not in cell_data.columns:
    #                 continue

    #             # Try to get base station lat/lon
    #             if "latitude" in row and "longitude" in row:
    #                 bs_lat, bs_lon = row["latitude"], row["longitude"]
    #             else:
    #                 # fallback, take mean of cell data
    #                 bs_lat, bs_lon = cell_data["latitude"].mean(), cell_data["longitude"].mean()

    #             plotter = SpatialKPIDensityPlot(
    #                 bs_lat=bs_lat, bs_lon=bs_lon,
    #                 azimuth=row.azimuth, beamwidth=row.beamwidth,
    #                 radius=100, grid_size=50,
    #                 data_points=cell_data,
    #                 lon_col="longitude",
    #                 lat_col="latitude",
    #                 kpi_col=kpi,
    #                 kpi_name=kpi.upper(),
    #                 kpi_range_dict=LTE_Ranges[kpi],
    #                 extent_km=16,
    #             )
    #             img = plotter.plot(out=None)  # returns openpyxl Image object
    #             self.plots[cell][kpi] = img

    def make_plots(self, kpi_list=None, *, pad=1.1, min_km=2.0, max_km=8.0):
        """Produce coloured PNGs with the new *Grid50Plot*."""
        kpi_list = kpi_list or list(LTE_Ranges.keys())
        self.plots = {}
        for cell in self.cells:
            df = self.all_data[self.all_data["siteid_cellid"] == cell]
            self.plots[cell] = {};  # even if empty (safe later)
            if df.empty: continue
            meta = self.overall_data[self.overall_data["siteid_cellid"] == cell].iloc[0]
            bs_lat = float(meta.get("latitude", df["latitude"].mean()))
            bs_lon = float(meta.get("longitude", df["longitude"].mean()))
            # extent heuristic: farthest GPS point distance
            dy = (df["latitude"] - bs_lat).abs().max() * 111_000
            dx = (df["longitude"]- bs_lon).abs().max() *  85_000
            win_km = np.clip(np.hypot(dx, dy)*pad/1_000, min_km, max_km)
            for kpi in kpi_list:
                if kpi not in df.columns: continue
                self.plots[cell][kpi] = Grid50Plot(
                    bs_lat=bs_lat, bs_lon=bs_lon,
                    azimuth=float(meta.get("azimuth", 0)),
                    beamwidth=float(meta.get("beamwidth", 60)),
                    data_points=df, kpi_col=kpi, kpi_name=kpi.upper(),
                    kpi_range_dict=LTE_Ranges[kpi], extent_km=round(float(win_km),1)
                ).plot(out=None)
        return self.plots

    def write_report_onepage(self, out_xlsx_path, site_name=None):

        wb = Workbook()
        ws = wb.active
        ws.title = "SSV 4G Report"
        ws.sheet_view.showGridLines = False

        # ---- Styles ----
        section_fill = PatternFill("solid", fgColor="3E82FC")
        table_header_fill = PatternFill("solid", fgColor="D1E6FA")
        table_band1 = PatternFill("solid", fgColor="F4F8FB")
        table_band2 = PatternFill("solid", fgColor="E9F1F7")
        border = Border(left=Side(style='thin', color='BBBBBB'),
                        right=Side(style='thin', color='BBBBBB'),
                        top=Side(style='thin', color='BBBBBB'),
                        bottom=Side(style='thin', color='BBBBBB'))
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_font = Font(bold=True, color="FFFFFF", size=16, name="Arial")
        subheader_font = Font(bold=True, color="3E82FC", size=13, name="Arial")
        kpiheader_font = Font(bold=True, color="000000", size=12, name="Arial")

        def section_header(ws, title, row, col=2, span=18):
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
            cell = ws.cell(row=row, column=col)
            cell.value = title
            cell.font = header_font
            cell.alignment = center
            cell.fill = section_fill
            return row + 2

        def write_table(ws, table, start_row, start_col=1, banded=True):
            for i, row_data in enumerate(table):
                for j, val in enumerate(row_data):
                    c = ws.cell(row=start_row + i, column=start_col + j, value=val)
                    c.border = border
                    c.alignment = center
                    if i == 0:
                        c.fill = table_header_fill
                        c.font = Font(bold=True, name="Arial")
                    elif banded and i % 2 == 0:
                        c.fill = table_band1
                    elif banded:
                        c.fill = table_band2
            return start_row + len(table)

        def auto_fit(ws):
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max(
                            dims.get(cell.column_letter, 0), len(str(cell.value))
                        )
            for col, value in dims.items():
                ws.column_dimensions[col].width = min(value + 4, 22)

        # ---- Start Writing ----
        cur_row = 2
        site_title = site_name if site_name is not None else getattr(self, "siteid", "Site")
        cur_row = section_header(ws, f"4G Site Report: {site_title}", cur_row, col=2, span=18)
        cur_row += 2

        # Overview heading & table
        ws.cell(row=cur_row, column=2, value="Overview").font = subheader_font
        ws.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row, end_column=8)
        cur_row += 1
        overview_tbl = self.overall_data.reset_index(drop=True)
        overview_tbl.columns = [str(c) for c in overview_tbl.columns]
        write_table(ws, [overview_tbl.columns.tolist()] + overview_tbl.astype(str).values.tolist(), cur_row, start_col=2)
        cur_row += len(overview_tbl) + 2

        # KPI heading & table
        ws.cell(row=cur_row, column=2, value="KPI").font = subheader_font
        ws.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row, end_column=8)
        cur_row += 1
        kpi_tbl = self.kpi.reset_index(drop=True)
        kpi_tbl.columns = [str(c) for c in kpi_tbl.columns]
        write_table(ws, [kpi_tbl.columns.tolist()] + kpi_tbl.astype(str).values.tolist(), cur_row, start_col=2)
        cur_row += len(kpi_tbl) + 2

        # Coverage and Performances heading
        cur_row = section_header(ws, "Coverage and Performances", cur_row, col=2, span=18)
        cur_row += 1

        # ---- Layout Config ----
        img_cols = 8         # columns for image
        img_rows = 17        # rows for image (big enough for map legend)
        table_cols = 7       # columns for table
        table_rows_height = 16
        gap = 3              # columns gap between blocks (image/table/chart)
        chart_width = 22
        chart_height = 13
        block_padding = 4    # empty rows after each KPI block

        # Adjust all column widths for visual consistency
        for c in range(2, 2+img_cols+table_cols+gap*2+8):
            ws.column_dimensions[get_column_letter(c)].width = 15

        # ---- For each cell, for each KPI ----
        for cell, kpis in self.tables.items():
            ws.cell(row=cur_row, column=2, value=f"Cell: {cell}").font = subheader_font
            ws.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row, end_column=8)
            cur_row += 1

            for kpi, table in kpis.items():
                ws.cell(row=cur_row, column=2, value=f"{kpi.upper()}").font = kpiheader_font
                ws.merge_cells(start_row=cur_row, start_column=2, end_row=cur_row, end_column=8)
                cur_row += 1

                # --- Image Block ---
                img_row = cur_row
                img_col = 2
                img_end_col = img_col + img_cols - 1
                img_end_row = img_row + img_rows - 1
                ws.merge_cells(start_row=img_row, start_column=img_col, end_row=img_end_row, end_column=img_end_col)
                # Calculate actual pixel size for the merged area
                total_width_pixels = sum(int(7 * ws.column_dimensions[get_column_letter(c)].width) for c in range(img_col, img_end_col+1))
                total_height_pixels = sum(int(1.33 * ws.row_dimensions[r].height if ws.row_dimensions[r].height else 16) for r in range(img_row, img_end_row+1))

                img = self.plots[cell][kpi] if cell in self.plots and kpi in self.plots[cell] else None
                if img:
                    img.width = total_width_pixels
                    img.height = total_height_pixels
                    img.anchor = f"{get_column_letter(img_col)}{img_row}"
                    ws.add_image(img)

                # --- Table Block ---
                table_col = img_end_col + gap
                table_row = img_row
                table_end_row = write_table(ws, table, start_row=table_row, start_col=table_col)
                for r in range(table_row, table_end_row):
                    ws.row_dimensions[r].height = table_rows_height

                # --- Chart Block ---
                chart_col = table_col + table_cols + gap
                header_row = table_row
                first_data_row = table_row + 1
                last_data_row = table_row + len(table) - 1
                data_col = table_col + 1
                cats_ref = Reference(ws, min_col=table_col, min_row=first_data_row, max_row=last_data_row)
                data_ref = Reference(ws, min_col=data_col, min_row=first_data_row, max_row=last_data_row)

                chart = BarChart()
                chart.add_data(data_ref, titles_from_data=False)
                chart.set_categories(cats_ref)
                chart.height = chart_height
                chart.width = chart_width
                chart.title = f"{kpi.upper()} Distribution"
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True  # Show only value
                chart.dataLabels.showPercent = False
                chart.dataLabels.showCatName = False
                chart.legend.position = "r"
                ws.add_chart(chart, f"{get_column_letter(chart_col)}{table_row}")

                # Add blank rows below each block to prevent overlapping
                cur_row = max(img_end_row, table_end_row) + block_padding

        auto_fit(ws)
        wb.save(out_xlsx_path)


    def __str__(self):
        return ", \n".join(f"{k}=\n{v}\n" for k, v in vars(self).items())

if __name__ == "__main__":
    from timeit import default_timer as timer

    start = timer()
    report = SSV4G(siteid="37233",task_date="2025-01-01")
    report.query_data()
    report.make_tables()
    report.make_plots()
    report.write_report_onepage(out_xlsx_path='C:/Users/erena/Desktop/Reporter/SSV_Report/SSV_Report.xlsx')
    print(str(report))
    end = timer()
    print(end - start) # Time in seconds, e.g. 5.38091952400282

